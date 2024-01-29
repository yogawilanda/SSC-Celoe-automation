import openpyxl
from tqdm import tqdm
import time

def find_empty_column(worksheet):
    next_empty_column = 1
    while worksheet.cell(row=1, column=next_empty_column).value is not None:
        next_empty_column += 1
        print(f"Column {openpyxl.utils.get_column_letter(next_empty_column - 1)} is not empty.")
    return next_empty_column

def insert_title(worksheet, column, title):
    worksheet.cell(row=1, column=column).value = title
    print(f"Title inserted in column {openpyxl.utils.get_column_letter(column)}.")
    
    
def insert_menu():
    # This is Menu 2 - Insertion - Convert it into separate file
    try:
        workbook = openpyxl.load_workbook(filename="aset/asetEdomGjl2223.xlsx")
        #insert worksheet
        input_worksheet = input("Enter the worksheet name: ")
        worksheet = workbook[input_worksheet]
        print(f"Filenya ketemu nih!\nNama sheet: {worksheet}.")
    
    except FileNotFoundError:
        print("File not found. Please check the file path.")
        return
    except KeyError:
        print("Worksheet not found. Please check the worksheet name.")
        return

    next_empty_column = find_empty_column(worksheet)

    # Prompt for user confirmation
    set_title_confirmation = input("Yes = Masukkan Judul Secara Manual\nNo = Judul Otomatis:\n ").lower().strip()

    if set_title_confirmation == "yes":
        # Prompt for the title
        title = input("Enter the title: ")
        insert_title(worksheet, next_empty_column, title)
    elif set_title_confirmation == "no":
        # Insert the title in the first row of the empty column
        insert_title(worksheet, next_empty_column, "Contoh")
    else:
        print("Invalid input. Please try again.")
        return
    
    # Insert data in the next rows of the same column, starting from the second row
    for i in range(2, worksheet.max_row + 1):  # Iterate over all rows in the worksheet
        previous_row_data = worksheet.cell(row=i - 1, column=next_empty_column).value
        if previous_row_data is not None:
            worksheet.cell(row=i, column=next_empty_column).value = worksheet.title
            print(f"Data inserted in column {openpyxl.utils.get_column_letter(next_empty_column)}.")
        else:
            break  # Break the loop if the previous row is empty

        # Save the modified Excel file
        workbook.save(filename="aset/asetEdomGjl2223.xlsx")
    
    else:
        print("Exiting program.")